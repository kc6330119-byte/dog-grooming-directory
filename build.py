#!/usr/bin/env python3
"""
Dog Groomer Locator - Static Site Generator

Fetches dog groomer listings from Airtable and generates a static HTML site.
Falls back to sample data if Airtable is not configured.
"""
import json
import os
import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
import markdown as md_lib
from jinja2 import Environment, FileSystemLoader
from markupsafe import Markup
from slugify import slugify

import config


def get_sample_data():
    """Return sample groomers for testing without Airtable."""
    return [
        {
            "name": "Pampered Paws Grooming Salon",
            "slug": "pampered-paws-grooming-salon-nashville",
            "description": "A full-service dog grooming salon offering breed-specific haircuts, hand-stripping, "
                           "deshedding treatments, nail grinding, and spa packages. Pampered Paws has served Nashville "
                           "pet owners for over 12 years with a focus on low-stress handling and personalized care for "
                           "every dog. Open six days a week with evening appointments available.",
            "address": "123 Main St",
            "city": "Nashville",
            "state": "Tennessee",
            "state_slug": "tennessee",
            "zip": "37201",
            "phone": "(615) 555-0100",
            "website_url": "",
            "google_maps_url": "",
            "photo_url": "",
            "hours": "Mon-Sat: 8am - 6pm",
            "services": ["Full Grooming", "Bath & Brush", "Nail Trimming", "Deshedding", "Teeth Brushing"],
            "specialties": ["Breed-Specific Styling", "Show Grooming"],
            "price_range": "$$",
            "type": "Full-Service Salon",
            "status": "Active",
            "date_added": "2025-01-01",
            "rating": 4.8,
            "review_count": 215,
            "last_modified": "2025-06-15",
        },
        {
            "name": "Bark & Clean Mobile Grooming",
            "slug": "bark-and-clean-mobile-grooming-austin",
            "description": "Bark & Clean brings professional grooming directly to your door in a fully equipped, "
                           "climate-controlled mobile grooming van. Specializing in anxious dogs and senior pets, they "
                           "offer one-on-one grooming sessions that reduce stress. Services include full grooming, "
                           "bath-only options, nail care, and puppy first groom introductions.",
            "address": "456 River Rd",
            "city": "Austin",
            "state": "Texas",
            "state_slug": "texas",
            "zip": "78701",
            "phone": "(512) 555-0200",
            "website_url": "",
            "google_maps_url": "",
            "photo_url": "",
            "hours": "Tue-Sun: 9am - 5pm",
            "services": ["Full Grooming", "Bath Only", "Nail Trimming", "Puppy First Groom"],
            "specialties": ["Mobile Grooming", "Anxious Dogs"],
            "price_range": "$$$",
            "type": "Mobile Grooming",
            "status": "Active",
            "date_added": "2025-01-02",
            "rating": 4.9,
            "review_count": 143,
            "last_modified": "2025-07-01",
        },
        {
            "name": "Suds & Scissors Pet Spa",
            "slug": "suds-and-scissors-pet-spa-chicago",
            "description": "Suds & Scissors is a neighborhood pet spa offering professional grooming for dogs and cats. "
                           "Their team of certified groomers handles all breeds and sizes, from Chihuahuas to Great Danes. "
                           "Services include spa baths, breed-standard cuts, creative grooming, and a self-service dog wash "
                           "station for DIY owners.",
            "address": "789 Oak Ave",
            "city": "Chicago",
            "state": "Illinois",
            "state_slug": "illinois",
            "zip": "60601",
            "phone": "(312) 555-0300",
            "website_url": "",
            "google_maps_url": "",
            "photo_url": "",
            "hours": "Mon-Fri: 7am - 7pm, Sat: 8am - 5pm",
            "services": ["Full Grooming", "Cat Grooming", "Self-Service Wash", "Creative Grooming"],
            "specialties": ["All Breeds", "Cat Grooming"],
            "price_range": "$$",
            "type": "Full-Service Salon",
            "status": "Active",
            "date_added": "2025-01-03",
            "rating": 4.6,
            "review_count": 89,
            "last_modified": "2025-08-10",
        },
    ]


def fetch_from_airtable():
    """Fetch groomers from Airtable API."""
    if not config.AIRTABLE_API_KEY or not config.AIRTABLE_BASE_ID:
        print("Airtable not configured. Using sample data.")
        return None

    try:
        from pyairtable import Api

        api = Api(config.AIRTABLE_API_KEY)
        table = api.table(config.AIRTABLE_BASE_ID, config.AIRTABLE_TABLE_NAME)
        records = table.all()

        groomers = []
        for record in records:
            fields = record.get("fields", {})

            if fields.get("Status") == "Draft":
                continue

            state_name = fields.get("State", "")
            groomer = {
                "_airtable_id": record.get("id", ""),
                "name": fields.get("Name", ""),
                "slug": slugify(fields.get("Name", "") + "-" + fields.get("City", "")),
                "description": RATING_CLAIM_RE.sub("", fields.get("Decription", fields.get("Description", ""))).strip(),
                "address": fields.get("Address", ""),
                "city": fields.get("City", ""),
                "state": state_name,
                "state_slug": slugify(state_name),
                "zip": fields.get("Zip", ""),
                "phone": fields.get("Phone", ""),
                "website_url": fields.get("Website URL", ""),
                "google_maps_url": fields.get("Google Maps URL", ""),
                "photo_url": fields.get("Photo URL", ""),
                "hours": fields.get("Hours", ""),
                "services": fields.get("Services", []),
                "specialties": fields.get("Specialties", []),
                "price_range": fields.get("Price Range", ""),
                "type": fields.get("Type", "Full-Service Salon"),
                "status": fields.get("Status", "Active"),
                "featured": fields.get("Featured", False),
                "date_added": fields.get("Date Added", ""),
                "rating": fields.get("Rating", 0),
                "review_count": fields.get("Review Count", 0),
                "latitude": fields.get("Latitude", ""),
                "longitude": fields.get("Longitude", ""),
                "last_modified": fields.get("Last Modified", fields.get("Date Added", "")),
            }
            groomers.append(groomer)

        print(f"Fetched {len(groomers)} groomers from Airtable.")

        # Disambiguate duplicate slugs by appending zip code
        slug_groups = {}
        for g in groomers:
            slug_groups.setdefault(g["slug"], []).append(g)
        for slug, group in slug_groups.items():
            if len(group) > 1:
                for g in group:
                    zip_code = g.get("zip", "").strip()
                    if zip_code:
                        g["slug"] = f"{slug}-{slugify(zip_code)}"

        return groomers

    except Exception as e:
        print(f"Error fetching from Airtable: {e}")
        return None


def clear_airtable_photo_url(airtable_id):
    """Clear the Photo URL field in Airtable for a groomer with an expired URL."""
    if not config.AIRTABLE_API_KEY or not config.AIRTABLE_BASE_ID or not airtable_id:
        return
    try:
        from pyairtable import Api
        api = Api(config.AIRTABLE_API_KEY)
        table = api.table(config.AIRTABLE_BASE_ID, config.AIRTABLE_TABLE_NAME)
        table.update(airtable_id, {"Photo URL": ""})
    except Exception as e:
        print(f"  Warning: could not clear Airtable photo_url for {airtable_id}: {e}")


# Third-party rating claims in body copy (every fact-description ended with
# "It holds a X-star rating from N Google reviews.") are a Google rating-snippet
# policy risk AND the site's #1 templated shingle (99% of indexable pages).
# The visible star badge in groomer.html keeps the rating with attribution;
# the body sentence is stripped at fetch. Same approach as the sister site.
PRICE_RE = re.compile(r"\$\d")  # any currency figure in a description

RATING_CLAIM_RE = re.compile(
    r"\s*It (?:holds|has|maintains) a [\d.]+-star rating(?: from [\d,]+ Google reviews?)?\.",
    re.IGNORECASE,
)


# Minimum listings a real build must produce. Airtable returns an EMPTY LIST
# (not None) when the Groomers table has no rows, so without this guard the build
# would succeed and publish a site with zero listings — silently replacing every
# live page. Fail loudly instead. Set ALLOW_SMALL_BUILD=true to override when a
# small or empty site is genuinely intended.
MIN_EXPECTED_LISTINGS = int(os.getenv("MIN_EXPECTED_LISTINGS", "500"))


def get_groomers():
    """Get groomers from Airtable or fall back to sample data."""
    groomers = fetch_from_airtable()
    if groomers is None:
        groomers = get_sample_data()
        print(f"Using {len(groomers)} sample groomers.")
        dedupe_slugs(groomers)
        return groomers

    allow_small = os.getenv("ALLOW_SMALL_BUILD", "").lower() in {"1", "true", "yes", "on"}
    if len(groomers) < MIN_EXPECTED_LISTINGS and not allow_small:
        raise SystemExit(
            f"\nABORTING BUILD: Airtable returned {len(groomers)} listings, "
            f"below the expected minimum of {MIN_EXPECTED_LISTINGS}.\n"
            "Publishing this would replace the live site with an empty directory.\n"
            "If the Airtable data was intentionally removed, the deployed site should be\n"
            "left as-is (do not redeploy). To restore, re-import the CSV backup in data/\n"
            "via upload_to_airtable.py. To build a genuinely small site anyway, set\n"
            "ALLOW_SMALL_BUILD=true.\n"
        )
    dedupe_slugs(groomers)
    return groomers


def dedupe_slugs(groomers):
    """Ensure every groomer has a stable, unique output slug."""
    used = set()
    for index, groomer in enumerate(groomers, start=1):
        base = (groomer.get("slug") or "").strip()
        if not base:
            base = slugify(f"{groomer.get('name', '')}-{groomer.get('city', '')}") or f"groomer-{index}"

        candidates = [base]
        zip_code = slugify(str(groomer.get("zip") or "").strip())
        if zip_code and not base.endswith(f"-{zip_code}"):
            candidates.append(f"{base}-{zip_code}")

        record_id = slugify(str(groomer.get("_airtable_id") or "").strip()[-6:])
        if record_id:
            candidates.append(f"{base}-{record_id}")

        candidate = next((c for c in candidates if c and c not in used), "")
        suffix = 2
        while not candidate:
            numbered = f"{base}-{suffix}"
            if numbered not in used:
                candidate = numbered
            suffix += 1

        groomer["slug"] = candidate
        used.add(candidate)


def fetch_blog_posts():
    """Fetch published blog posts from Airtable."""
    if not config.AIRTABLE_API_KEY or not config.AIRTABLE_BASE_ID:
        return []

    try:
        from pyairtable import Api

        api = Api(config.AIRTABLE_API_KEY)
        table = api.table(config.AIRTABLE_BASE_ID, config.AIRTABLE_BLOG_TABLE_NAME)
        records = table.all()

        today = datetime.now().strftime("%Y-%m-%d")
        posts = []
        for record in records:
            fields = record.get("fields", {})

            if fields.get("Status") != "Published":
                continue

            # Skip posts with a future publish date
            publish_date = fields.get("Publish Date", "")
            if publish_date and publish_date > today:
                continue

            title = fields.get("Title", "")

            # Author attribution resolves strictly through config.AUTHORS.
            # Raw Airtable "Author"/"Author Bio" text is never rendered, so
            # retired bylines in stale records cannot resurface. Unknown keys
            # or names fall back to the editorial team.
            author_key = fields.get("Author Key", "")
            if author_key not in config.AUTHORS:
                airtable_author = (fields.get("Author", "") or "").strip().lower()
                author_key = config.DEFAULT_AUTHOR_KEY
                for key, info in config.AUTHORS.items():
                    if info["name"].lower() == airtable_author:
                        author_key = key
                        break
            author_info = config.AUTHORS[author_key]
            author_name = author_info["name"]
            author_credentials = author_info["credentials"]
            author_bio_text = author_info["bio"]

            post = {
                "title": title,
                "slug": (fields.get("Slug", "") or slugify(title)).strip(),
                "content": fields.get("Content", ""),
                "excerpt": fields.get("Excerpt", ""),
                "author": author_name,
                "author_key": author_key,
                "author_credentials": author_credentials,
                "author_bio": author_bio_text,
                "publish_date": fields.get("Publish Date", ""),
                "modified_date": fields.get("Modified Date", fields.get("Publish Date", "")),
                "featured_image": fields.get("Featured Image", ""),
                "meta_description": fields.get("Meta Description", ""),
                "status": fields.get("Status", "Published"),
                "featured": fields.get("Featured", False),
            }
            posts.append(post)

        posts.sort(key=lambda x: x.get("publish_date", ""), reverse=True)
        print(f"Fetched {len(posts)} blog posts from Airtable.")
        return posts

    except Exception as e:
        print(f"Note: Could not fetch blog posts ({e})")
        return []


def setup_output_directory():
    """Create clean output directory."""
    if config.OUTPUT_DIR.exists():
        shutil.rmtree(config.OUTPUT_DIR)

    config.OUTPUT_DIR.mkdir(parents=True)
    (config.OUTPUT_DIR / "state").mkdir()
    (config.OUTPUT_DIR / "groomer").mkdir()
    (config.OUTPUT_DIR / "category").mkdir()
    (config.OUTPUT_DIR / "blog").mkdir()

    # Copy static files
    if config.STATIC_DIR.exists():
        shutil.copytree(config.STATIC_DIR, config.OUTPUT_DIR / "static")


def create_jinja_env():
    """Create Jinja2 environment with custom filters."""
    env = Environment(
        loader=FileSystemLoader(config.TEMPLATES_DIR),
        autoescape=True
    )

    def format_date(date_str):
        if not date_str:
            return ""
        try:
            dt = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
            return dt.strftime("%B ") + str(dt.day) + dt.strftime(", %Y")
        except (ValueError, TypeError):
            return date_str

    env.filters["slugify"] = slugify
    env.filters["tojson"] = lambda v: json.dumps(v, ensure_ascii=False)
    env.filters["markdown"] = lambda text: Markup(md_lib.markdown(text or "", extensions=["extra", "nl2br"]))
    env.filters["format_date"] = format_date

    env.globals["site_name"] = config.SITE_NAME
    env.globals["site_url"] = config.SITE_URL
    env.globals["site_description"] = config.SITE_DESCRIPTION
    env.globals["categories"] = config.CATEGORIES
    env.globals["us_states"] = config.US_STATES
    env.globals["current_year"] = datetime.now().year
    env.globals["ga_measurement_id"] = config.GA_MEASUREMENT_ID
    env.globals["adsense_enabled"] = config.ADSENSE_ENABLED
    env.globals["adsense_client_id"] = config.ADSENSE_CLIENT_ID
    env.globals["default_og_image"] = config.DEFAULT_OG_IMAGE
    env.globals["authors"] = config.AUTHORS

    return env


def group_by_state(groomers):
    """Group groomers by state slug."""
    grouped = {}
    for g in groomers:
        state_slug = g.get("state_slug", "")
        if state_slug:
            grouped.setdefault(state_slug, []).append(g)
    return grouped


def group_by_city(groomers):
    """Group groomers by (state_slug, city) for city pages."""
    grouped = defaultdict(list)
    for g in groomers:
        state_slug = g.get("state_slug", "")
        city = g.get("city", "").strip()
        if state_slug and city:
            city_slug = slugify(city)
            grouped[(state_slug, city, city_slug)].append(g)
    return grouped


def truncate_at_word(value, max_length):
    """Trim text cleanly without splitting the final word."""
    text = " ".join(str(value or "").split())
    if len(text) <= max_length:
        return text
    trimmed = text[:max_length - 3].rsplit(" ", 1)[0].rstrip(" ,;:-|")
    return f"{trimmed}..." if trimmed else f"{text[:max_length - 3]}..."


def seo_title(primary, secondary="", max_length=62):
    """Build a compact, keyword-first page title."""
    primary = " ".join(str(primary or "").split())
    secondary = " ".join(str(secondary or "").split())
    if secondary:
        candidate = f"{primary} | {secondary}"
        if len(candidate) <= max_length:
            return candidate
    if len(primary) <= max_length:
        return primary
    return truncate_at_word(primary, max_length)


CATEGORY_SEO_TITLES = {
    "full-service": "Full-Service Dog Groomers Near Me",
    "mobile-grooming": "Mobile Dog Groomers Near Me",
    "breed-specific": "Breed-Specific Dog Groomers Near Me",
    "cat-grooming": "Cat Grooming Services Near Me",
    "puppy-first-groom": "Puppy Groomers Near Me",
    "senior-special-needs": "Senior Dog & Special Needs Groomers Near Me",
    "self-service-wash": "Self-Service Dog Wash Near Me",
    "affordable": "Affordable Dog Groomers Near Me",
}


CATEGORY_FILTERS = {
    "full-service": lambda g: g.get("type") in ["Full-Service Salon", "Full-Service Grooming"],
    "mobile-grooming": lambda g: g.get("type") == "Mobile Grooming" or "Mobile Grooming" in g.get("specialties", []),
    "breed-specific": lambda g: "Breed-Specific Styling" in g.get("specialties", []) or "Show Grooming" in g.get("specialties", []),
    "cat-grooming": lambda g: "Cat Grooming" in g.get("services", []) or "Cat Grooming" in g.get("specialties", []),
    "puppy-first-groom": lambda g: "Puppy First Groom" in g.get("services", []) or "Puppy First Groom" in g.get("specialties", []),
    "senior-special-needs": lambda g: "Senior Dogs" in g.get("specialties", []) or "Special Needs" in g.get("specialties", []),
    "self-service-wash": lambda g: "Self-Service Wash" in g.get("services", []) or g.get("type") == "Self-Service Dog Wash",
    "affordable": lambda g: g.get("price_range") in ["$", "$$"],
}


def category_groomers_for(groomers, category):
    """Return quality-passing listings that match a public category."""
    filter_fn = CATEGORY_FILTERS.get(category.get("slug", ""), lambda g: True)
    return [g for g in groomers if filter_fn(g)]


def category_seo_title(category):
    """Return a natural category title for local-intent searches."""
    return CATEGORY_SEO_TITLES.get(
        category.get("slug", ""),
        f"{category.get('name', 'Dog Groomers')} Near Me",
    )


def groomer_page_title(groomer):
    """Use business name plus local service context for detail-page titles."""
    name = groomer.get("name", "Dog Groomer")
    city = groomer.get("city", "")
    state = groomer.get("state", "")
    location = ", ".join(part for part in [city, state] if part)
    has_grooming_context = re.search(
        r"\b(groom|grooming|groomer|salon|spa|pet|pooch|puppy|dog|hound|canine|k9|k-9)\b",
        str(name).lower(),
    )
    service_label = "" if has_grooming_context else " Dog Groomer"
    primary = f"{name}{service_label} in {location}" if location else f"{name}{service_label}"
    return seo_title(primary, "Reviews & Info", max_length=70)


def generate_dynamic_meta_description(page_type, context):
    """Generate unique meta description from actual page content — never formulaic.

    Must-have #1: Unique dynamic meta descriptions derived from editorial content.
    """
    if page_type == "groomer":
        name = str(context.get("name", "")).strip()
        city = str(context.get("city", "")).strip()
        state = str(context.get("state", "")).strip()
        desc = str(context.get("description", "")).strip()

        if desc and len(desc) >= 50:
            # Descriptions already lead with the business name, so they're
            # unique on their own. Only prepend a name prefix when the
            # description doesn't already start with the business name.
            if name and not desc.lower().startswith(name.lower()):
                prefix = f"{name} is a dog groomer in {city}, {state}" if name and city else name
                candidate = f"{prefix} — {desc}" if prefix else desc
            else:
                candidate = desc
            return truncate_at_word(candidate, 155)

        # Fallback: structured meta from business data
        if name and city and state:
            return truncate_at_word(
                f"{name} is a dog groomer in {city}, {state}. View services, hours, reviews, and contact info.",
                155,
            )
        return ""

    elif page_type == "state":
        state = str(context.get("state", "")).strip()
        count = context.get("count")
        if state:
            count_text = f"{count} " if count else ""
            return truncate_at_word(
                f"Find {count_text}dog groomers in {state}. Browse local salons, mobile groomers, hours, services, reviews, and state-specific grooming tips.",
                155,
            )
        state_desc = str(context.get("state_description", "")).strip()
        if state_desc:
            return truncate_at_word(state_desc, 155)
        return ""

    elif page_type == "city":
        city = str(context.get("city", "")).strip()
        state = str(context.get("state", "")).strip()
        count = context.get("count")
        if city and state:
            count_text = f"{count} " if count else ""
            return truncate_at_word(
                f"Find {count_text}dog groomers near you in {city}, {state}. Compare local grooming services, hours, reviews, and contact info.",
                155,
            )
        return ""

    elif page_type == "category":
        category_name = str(context.get("category_name", "")).strip()
        if category_name:
            category_label = re.sub(r"\s+near me$", "", category_name, flags=re.IGNORECASE)
            return truncate_at_word(
                f"Find {category_label.lower()} near you. Browse quality-checked dog groomers by state with services, hours, reviews, and contact details.",
                155,
            )
        intro = str(context.get("intro", "")).strip()
        if intro:
            return truncate_at_word(intro, 155)
        return ""

    elif page_type == "post":
        meta = str(context.get("meta_description", "")).strip()
        if meta:
            return meta[:155]
        excerpt = str(context.get("excerpt", "")).strip()
        if excerpt:
            return truncate_at_word(excerpt, 155)
        return ""

    return ""


def load_gsc_protected_slugs():
    """Load groomer slugs that are currently earning Google traffic.

    These are grandfathered through description_quality_check so the new
    strict gate cannot regress any URL that's already ranking. Refresh by
    running scripts/extract_gsc_protected.py against a fresh GSC export.
    """
    path = config.BASE_DIR / "data" / "gsc_protected_urls.txt"
    if not path.exists():
        return set()
    return {
        line.strip()
        for line in path.read_text().splitlines()
        if line.strip() and not line.startswith("#")
    }


_GSC_PROTECTED = load_gsc_protected_slugs()


def description_start_key(description):
    """Normalize a description start so repeated boilerplate can be detected."""
    words = re.findall(r"[a-z0-9]+", str(description or "").lower())
    if len(words) < config.REPETITIVE_DESCRIPTION_START_WORDS:
        return ""
    return " ".join(words[:config.REPETITIVE_DESCRIPTION_START_WORDS])


def description_quality_check(groomer):
    """Decide whether a groomer listing has enough genuine content to be indexed.

    Returns (passes, reason). Description is the primary gate — its length and
    content quality matter more than whether other data fields are populated.
    Hours/phone/services are demoted to secondary signals that can rescue a
    borderline listing.

    GSC-protected slugs (currently earning Google traffic) are grandfathered:
    they pass regardless of the new gate, so the rewrite cannot regress
    pages that are already ranking. After the 60-day window the protection
    list should be re-extracted from a fresh GSC report.
    """
    slug = (groomer.get("slug") or "").lower()
    desc = str(groomer.get("description") or "").strip()
    desc_lower = desc.lower()

    if not desc or desc_lower == "nan":
        return False, "no description"
    for term in config.JUNK_DESCRIPTION_TERMS:
        if term in desc_lower:
            return False, f"junk pattern: {term!r}"
    if not any(v in desc_lower for v in config.GROOMING_VOCAB):
        return False, "no grooming vocabulary"
    if slug in _GSC_PROTECTED:
        return True, "gsc-protected"
    if len(desc) < config.MIN_DESCRIPTION_LENGTH:
        return False, f"description too short ({len(desc)} < {config.MIN_DESCRIPTION_LENGTH})"

    # Description passed the primary gate. Require at least one additional
    # signal so we don't index a listing with rich text but no contact info.
    hours_ok = bool(str(groomer.get("hours") or "").strip()) and \
               str(groomer.get("hours") or "").strip().lower() != "nan"
    services = groomer.get("services") or []
    services_ok = isinstance(services, list) and len(services) > 0
    phone_ok = bool(str(groomer.get("phone") or "").strip())
    if not (hours_ok or services_ok or phone_ok):
        return False, "no contact/service signals"

    return True, "ok"


def annotate_listing_quality(groomers):
    """Attach content-quality decisions to groomers for build reuse."""
    starts = defaultdict(int)
    for groomer in groomers:
        key = description_start_key(groomer.get("description", ""))
        if key:
            starts[key] += 1

    for groomer in groomers:
        passes, reason = description_quality_check(groomer)
        key = description_start_key(groomer.get("description", ""))
        if (
            passes
            and key
            and starts[key] > config.REPETITIVE_DESCRIPTION_MAX_REPEATS
            and (groomer.get("slug") or "").lower() not in _GSC_PROTECTED
        ):
            passes = False
            reason = f"repetitive description start ({starts[key]} matches)"

        groomer["_quality_passes"] = passes
        groomer["_quality_reason"] = reason


def is_quality_listing(groomer):
    """Return true when a listing is suitable for public discovery surfaces."""
    if "_quality_passes" not in groomer:
        passes, reason = description_quality_check(groomer)
        groomer["_quality_passes"] = passes
        groomer["_quality_reason"] = reason
    return bool(groomer.get("_quality_passes"))


def quality_reason(groomer):
    if "_quality_reason" not in groomer:
        is_quality_listing(groomer)
    return groomer.get("_quality_reason", "")


def public_groomers(groomers):
    """Only quality-passing listings should appear in normal browse/search flows."""
    return [g for g in groomers if is_quality_listing(g)]


def is_thin_listing(groomer):
    """Backward-compatible wrapper around description_quality_check."""
    return not is_quality_listing(groomer)


def build_homepage(env, groomers, posts):
    """Build the homepage."""
    template = env.get_template("index.html")

    featured = [g for g in groomers if g.get("featured")][:config.FEATURED_COUNT]
    if not featured:
        featured = groomers[:config.FEATURED_COUNT]

    recent = sorted(groomers, key=lambda x: x.get("date_added", ""), reverse=True)[:config.RECENT_COUNT]

    by_state = group_by_state(groomers)
    state_counts = {s: len(v) for s, v in by_state.items()}

    featured_post = next((p for p in posts if p.get("featured")), None)
    recent_posts = [p for p in posts if p is not featured_post][:3]

    html = template.render(
        featured_groomers=featured,
        recent_groomers=recent,
        all_groomers=groomers,
        state_counts=state_counts,
        total_count=len(groomers),
        posts=posts,
        featured_post=featured_post,
        recent_posts=recent_posts,
        page_title=config.DEFAULT_META_TITLE,
        meta_description=config.DEFAULT_META_DESCRIPTION,
        request_path="/",
    )

    output_path = config.OUTPUT_DIR / "index.html"
    output_path.write_text(html)
    print(f"Built: index.html ({len(groomers)} total groomers)")


def build_state_pages(env, groomers):
    """Build one page per US state."""
    template = env.get_template("state.html")
    grouped = group_by_state(groomers)

    for state in config.US_STATES:
        state_groomers = grouped.get(state["slug"], [])
        state_groomers.sort(key=lambda x: x.get("city", ""))

        thin_state = len(state_groomers) < config.MIN_STATE_LISTINGS_FOR_INDEX

        meta_desc = generate_dynamic_meta_description("state", {
            "state": state["name"],
            "count": len(state_groomers),
            "state_description": state["description"],
        })

        html = template.render(
            state=state,
            groomers=state_groomers,
            page_title=seo_title(f"Dog Groomers in {state['name']}", "Local Salons Near You"),
            meta_description=meta_desc,
            request_path=f"/state/{state['slug']}",
            noindex=thin_state,
        )

        output_path = config.OUTPUT_DIR / "state" / f"{state['slug']}.html"
        output_path.write_text(html)
        print(f"Built: state/{state['slug']}.html ({len(state_groomers)} groomers)")


def build_city_pages(env, groomers):
    """Build city-level pages.

    Pages are still emitted for cities with >= 2 listings so internal links
    from groomer detail pages resolve. Cities below MIN_CITY_LISTINGS_FOR_INDEX
    are noindexed and excluded from the sitemap — this is the audit-D fix:
    the prior implementation injected a templated 30-word intro on every city
    page, producing 749 indexed near-duplicates that AdSense classifies as
    doorway pages. The intro is now removed entirely; per-city editorial can
    be added later via an Airtable field if real local content is written.
    """
    template = env.get_template("city.html")
    city_groups = group_by_city(groomers)

    city_count = 0
    indexed_count = 0
    for (state_slug, city, city_slug), city_groomers in sorted(city_groups.items()):
        if len(city_groomers) < 2:
            continue

        state_dir = config.OUTPUT_DIR / "state" / state_slug
        state_dir.mkdir(parents=True, exist_ok=True)

        state_name = city_groomers[0].get("state", "")
        thin_city = len(city_groomers) < config.MIN_CITY_LISTINGS_FOR_INDEX
        if not thin_city:
            indexed_count += 1

        meta_desc = generate_dynamic_meta_description("city", {
            "city": city,
            "state": state_name,
            "count": len(city_groomers),
        })

        state_info = next((s for s in config.US_STATES if s["slug"] == state_slug), None)

        html = template.render(
            city=city,
            city_slug=city_slug,
            state=state_info or {"name": state_name, "slug": state_slug},
            groomers=city_groomers,
            city_intro="",  # audit-D: no templated intro
            page_title=seo_title(f"Dog Groomers in {city}, {state_name}", "Salons Near You"),
            meta_description=meta_desc,
            request_path=f"/state/{state_slug}/{city_slug}",
            noindex=thin_city,
        )

        output_path = state_dir / f"{city_slug}.html"
        output_path.write_text(html)
        city_count += 1

    print(f"Built: {city_count} city pages ({indexed_count} indexed, {city_count - indexed_count} noindexed)")


_GUIDE_LINKS_COMMON = [
    ("How to Choose a Dog Groomer Near You", "/blog/how-to-choose-a-dog-groomer-near-you"),
    ("What to Expect at a Grooming Appointment", "/blog/what-happens-at-a-dog-grooming-appointment-step-by-step"),
    ("Dog Grooming Prices Explained", "/blog/dog-grooming-prices-explained"),
]
_GUIDE_LINKS_BY_TYPE = {
    "Mobile Grooming": [
        ("Mobile vs. Salon Grooming: Which Is Right?", "/blog/mobile-vs-salon-dog-grooming-which-is-right"),
        ("How to Find a Mobile Dog Groomer", "/blog/how-to-find-mobile-dog-groomer"),
    ],
    "Self-Service Wash": [
        ("How to Groom Your Dog at Home", "/blog/how-to-groom-dog-at-home"),
        ("Essential Dog Grooming Tools", "/blog/essential-dog-grooming-tools-guide"),
    ],
    "Pet Spa": [
        ("Between-Groom Maintenance Tips", "/blog/between-groom-maintenance-tips"),
    ],
}


def groomer_localbusiness_jsonld(groomer):
    """Assemble valid LocalBusiness JSON-LD. Conditional fields are only included
    when present so we never emit empty/invalid values. openingHours is omitted:
    the source is a freeform human-readable string, not the schema.org format, so
    emitting it would be invalid markup."""
    data = {
        "@context": "https://schema.org",
        "@type": "LocalBusiness",
        "name": groomer.get("name"),
        "description": groomer.get("description"),
        "address": {
            "@type": "PostalAddress",
            "streetAddress": groomer.get("address"),
            "addressLocality": groomer.get("city"),
            "addressRegion": groomer.get("state"),
            "postalCode": groomer.get("zip"),
            "addressCountry": "US",
        },
    }
    if groomer.get("photo_url"):
        data["image"] = groomer["photo_url"]
    if groomer.get("phone"):
        data["telephone"] = groomer["phone"]
    if groomer.get("website_url"):
        data["url"] = groomer["website_url"]
    if groomer.get("google_maps_url"):
        data["hasMap"] = groomer["google_maps_url"]
    price_range = str(groomer.get("price_range") or "").strip()
    if price_range:
        data["priceRange"] = price_range
    return data


def grooming_guide_links(groomer):
    """Contextual internal links to editorial content, varied by listing type so
    every page links into the site's real guides rather than carrying identical
    boilerplate."""
    links = list(_GUIDE_LINKS_BY_TYPE.get(groomer.get("type", ""), []))
    for item in _GUIDE_LINKS_COMMON:
        if len(links) >= 4:
            break
        links.append(item)
    links.append(("Dog Grooming Guide by Breed", "/dog-grooming-guide"))
    return [{"label": l, "url": u} for l, u in links]


def build_groomer_pages(env, groomers, public_listing_pool):
    """Build individual groomer detail pages."""
    template = env.get_template("groomer.html")
    gone_count = 0

    for groomer in groomers:
        related = [g for g in public_listing_pool if g["slug"] != groomer["slug"]
                   and g.get("state_slug") == groomer.get("state_slug")][:4]

        # Listings that fail the quality gate get NO page at all. They used to render a
        # "being reviewed for accuracy" stub, which (a) claimed an editorial review process
        # the site does not perform — AdSense P4 forbids claiming services not provided —
        # and (b) was a ~30-word under-construction screen, barred by P1. There is no data
        # to expand them from (Services 17.2% filled, Specialties 11.4%), so per P2's
        # expand-or-consolidate rule they are consolidated away: the URL returns 410 Gone
        # (see build_netlify_redirects). This reverses automatically if a listing's data
        # later passes the gate.
        if is_thin_listing(groomer):
            gone_count += 1
            continue

        meta_desc = generate_dynamic_meta_description("groomer", {
            "description": groomer.get("description", ""),
            "name": groomer.get("name", ""),
            "city": groomer.get("city", ""),
            "state": groomer.get("state", ""),
        })

        html = template.render(
            groomer=groomer,
            related_groomers=related,
            guide_links=grooming_guide_links(groomer),
            groomer_jsonld=groomer_localbusiness_jsonld(groomer),
            page_title=groomer_page_title(groomer),
            meta_description=meta_desc,
            request_path=f"/groomer/{groomer['slug']}",
            noindex=False,       # gate-failing listings no longer reach this point
            suppress_ads=False,
            quality_reason=quality_reason(groomer),
            # Prices come from the business's own website and go stale. Disclose provenance
            # and a capture date rather than presenting a figure as current fact.
            description_has_price=bool(PRICE_RE.search(groomer.get("description", "") or "")),
            description_sourced_on=config.SITE_DATA_CHECKED_ON,
        )

        output_path = config.OUTPUT_DIR / "groomer" / f"{groomer['slug']}.html"
        output_path.write_text(html)

    print(f"Built: {len(groomers) - gone_count} groomer pages "
          f"({gone_count} gate-failing listings omitted -> 410 Gone)")


def build_category_pages(env, groomers):
    """Build service category pages."""
    template = env.get_template("category.html")

    for category in config.CATEGORIES:
        category_groomers = category_groomers_for(groomers, category)
        preview_groomers = category_groomers[:config.CATEGORY_PREVIEW_LIMIT]

        state_counts = {}
        for g in category_groomers:
            s = g.get("state", "")
            if s:
                slug = g.get("state_slug", "")
                state_counts[s] = {"name": s, "slug": slug, "count": state_counts.get(s, {}).get("count", 0) + 1}
        state_list = sorted(state_counts.values(), key=lambda x: (-x["count"], x["name"]))

        meta_desc = generate_dynamic_meta_description("category", {
            "intro": category["intro"],
            "category_name": category_seo_title(category),
        })

        html = template.render(
            category=category,
            groomers=preview_groomers,
            total_count=len(category_groomers),
            showing_limited=len(category_groomers) > len(preview_groomers),
            state_list=state_list,
            page_title=seo_title(category_seo_title(category), "Local Options"),
            meta_description=meta_desc,
            request_path=f"/category/{category['slug']}",
            noindex=len(category_groomers) == 0,
            suppress_ads=len(category_groomers) == 0,
        )

        output_path = config.OUTPUT_DIR / "category" / f"{category['slug']}.html"
        output_path.write_text(html)
        print(f"Built: category/{category['slug']}.html ({len(category_groomers)} groomers)")


# --- Editorial internal-linking / topic clusters -----------------------------
# The post template previously picked the FIRST 3 entries of all_posts as
# "related", so the same 3 posts collected every internal link and 32 of 36
# posts received none. Relatedness is now scored on shared title terms, with a
# coverage pass guaranteeing every post is linked from at least one other post.

_STOPWORDS = {
    "the", "a", "an", "and", "or", "for", "your", "you", "how", "what", "why",
    "to", "of", "in", "at", "on", "with", "guide", "dog", "dogs", "grooming",
    "groom", "complete", "everything", "need", "know", "that", "this", "is",
    "it", "its", "about", "when", "from", "their", "them", "should", "does",
}


def _post_ref(post):
    """Minimal, acyclic representation of a post for use in related-post lists."""
    return {
        "slug": post.get("slug", ""),
        "title": post.get("title", ""),
        "excerpt": post.get("excerpt", ""),
        "publish_date": post.get("publish_date", ""),
    }


def _title_terms(post):
    words = re.findall(r"[a-z]+", str(post.get("title", "")).lower())
    return {w for w in words if len(w) > 2 and w not in _STOPWORDS}


def compute_related_posts(posts, k=3):
    """Attach `related` (k topically-closest posts) to every post.

    Returns nothing; mutates each post dict. Scoring is term overlap on titles,
    tie-broken by recency, so breed guides cluster with breed guides and
    how-to/cost posts cluster with each other.
    """
    terms = {p["slug"]: _title_terms(p) for p in posts if p.get("slug")}
    inbound = defaultdict(int)

    for post in posts:
        slug = post.get("slug")
        if not slug:
            continue
        mine = terms[slug]
        scored = []
        for other in posts:
            oslug = other.get("slug")
            if not oslug or oslug == slug:
                continue
            overlap = len(mine & terms[oslug])
            scored.append((overlap, str(other.get("publish_date", "")), oslug, other))
        scored.sort(key=lambda x: (-x[0], x[1]), reverse=False)
        scored.sort(key=lambda x: -x[0])
        # Store lightweight refs, not the post objects themselves: a post's
        # related list would otherwise point back at posts that point at it,
        # and the sitemap datestore JSON-serializes post data (circular ref).
        post["related"] = [_post_ref(o) for _, _, _, o in scored[:k]]
        for o in post["related"]:
            inbound[o["slug"]] += 1

    # Coverage pass: any post nobody links to gets swapped into the related list
    # of the post it is most similar to (replacing that post's weakest pick).
    for post in posts:
        slug = post.get("slug")
        if not slug or inbound[slug]:
            continue
        best, best_score = None, -1
        for other in posts:
            oslug = other.get("slug")
            if not oslug or oslug == slug:
                continue
            score = len(terms[slug] & terms[oslug])
            if score > best_score:
                best, best_score = other, score
        if best and best.get("related"):
            dropped = best["related"].pop()
            inbound[dropped["slug"]] -= 1
            best["related"].append(_post_ref(post))
            inbound[slug] += 1


def build_blog_page(env, posts):
    """Build the blog listing page."""
    template = env.get_template("blog.html")
    html = template.render(
        posts=posts,
        page_title=f"Blog - {config.SITE_NAME}",
        meta_description="Tips, guides, and expert advice on dog grooming, coat care, and finding the right groomer for your pet.",
        request_path="/blog",
    )
    output_path = config.OUTPUT_DIR / "blog.html"
    output_path.write_text(html)
    print(f"Built: blog.html ({len(posts)} posts)")


def build_post_pages(env, posts):
    """Build individual blog post pages with BlogPosting JSON-LD and author bios."""
    template = env.get_template("post.html")
    compute_related_posts(posts)

    for post in posts:
        if not post.get("slug"):
            continue

        meta_desc = generate_dynamic_meta_description("post", {
            "meta_description": post.get("meta_description", ""),
            "excerpt": post.get("excerpt", ""),
        })

        html = template.render(
            post=post,
            all_posts=posts,
            related_posts=post.get("related", []),
            # H1 keeps the full editorial title; the <title> tag is capped so it
            # is not truncated in the SERP (27 of 36 were over the display limit).
            page_title=seo_title(post["title"], config.SITE_NAME),
            meta_description=meta_desc,
            request_path=f"/blog/{post['slug']}",
        )
        output_path = config.OUTPUT_DIR / "blog" / f"{post['slug']}.html"
        output_path.write_text(html)
        print(f"Built: blog/{post['slug']}.html")


def build_search_index(groomers):
    """Generate search-index.json for client-side search."""
    index = [
        {"name": g["name"], "city": g.get("city", ""), "state": g.get("state", ""), "slug": g["slug"]}
        for g in groomers if g.get("name") and g.get("slug")
    ]
    output_path = config.OUTPUT_DIR / "search-index.json"
    with open(output_path, "w") as f:
        json.dump(index, f, ensure_ascii=False)
    print(f"Built: search-index.json ({len(index)} groomers)")


def _stable_lastmods(entries, data_by_url, today):
    """Per-URL lastmod from a committed content-hash datestore (sitemap_lastmod.json).

    The Airtable "Last Modified" column froze at the 2026-03/04 import, so the
    sitemap told Google "nothing changed" straight through the 5/28 de-spin —
    suppressing the recrawl the remediation depended on. This store bumps a
    URL's date only when its content actually changes: groomer/blog URLs hash
    canonical JSON of their source data (environment-independent, same dates on
    Netlify and local); other URLs hash their rendered file. Committed to git.
    """
    import hashlib
    import json as _json
    store_path = config.BASE_DIR / "sitemap_lastmod.json"
    try:
        store = _json.loads(store_path.read_text())
    except (OSError, ValueError):
        store = {}
    dates = {}
    new_store = {}
    for url, _priority, _default in entries:
        rec = store.get(url) or {}
        data = data_by_url.get(url)
        if data is not None:
            basis = _json.dumps(data, sort_keys=True, ensure_ascii=True, default=str)
            digest = hashlib.sha256(basis.encode("utf-8")).hexdigest()[:16]
        else:
            rel = url.replace(config.SITE_URL, "").strip("/")
            page = config.OUTPUT_DIR / (f"{rel}.html" if rel else "index.html")
            try:
                digest = hashlib.sha256(page.read_bytes()).hexdigest()[:16]
            except OSError:
                digest = rec.get("hash", "")
        if digest and digest == rec.get("hash"):
            date = rec.get("date") or today
        elif not digest:
            date = rec.get("date") or today
        else:
            date = today
        dates[url] = date
        new_store[url] = {"hash": digest, "date": date}
    store_path.write_text(_json.dumps(new_store, indent=0, sort_keys=True) + "\n")
    return dates


def build_sitemap(groomers, posts, breeds=None):
    """Generate sitemap.xml with per-page lastmod dates.

    Must-have #8: Sitemap with actual lastmod dates, not just today's build date.
    """
    today = datetime.now().strftime("%Y-%m-%d")

    entries = [
        (f"{config.SITE_URL}/", "1.0", today),
        (f"{config.SITE_URL}/blog", "0.8", today),
        (f"{config.SITE_URL}/about", "0.5", today),
        (f"{config.SITE_URL}/contact", "0.4", today),
        (f"{config.SITE_URL}/disclaimer", "0.3", today),
        (f"{config.SITE_URL}/editorial-standards", "0.3", today),
        (f"{config.SITE_URL}/privacy", "0.3", today),
        (f"{config.SITE_URL}/terms", "0.3", today),
    ]

    if breeds:
        entries.append((f"{config.SITE_URL}/dog-grooming-guide", "0.8", today))

    grouped = group_by_state(groomers)
    for state in config.US_STATES:
        state_groomers = grouped.get(state["slug"], [])
        if len(state_groomers) >= config.MIN_STATE_LISTINGS_FOR_INDEX:
            # Use most recent listing modification date for state page lastmod
            state_lastmod = max(
                (g.get("last_modified", g.get("date_added", today)) or today for g in state_groomers),
                default=today,
            )
            entries.append((f"{config.SITE_URL}/state/{state['slug']}", "0.8", state_lastmod))

    # City pages — only include if above the noindex threshold. The page itself
    # is still served (so internal links work), but Google should not be told
    # about a noindexed URL via sitemap.
    city_groups = group_by_city(groomers)
    for (state_slug, city, city_slug), city_groomers in sorted(city_groups.items()):
        if len(city_groomers) >= config.MIN_CITY_LISTINGS_FOR_INDEX:
            city_lastmod = max(
                (g.get("last_modified", g.get("date_added", today)) or today for g in city_groomers),
                default=today,
            )
            entries.append((f"{config.SITE_URL}/state/{state_slug}/{city_slug}", "0.7", city_lastmod))

    for category in config.CATEGORIES:
        if category_groomers_for(groomers, category):
            entries.append((f"{config.SITE_URL}/category/{category['slug']}", "0.7", today))

    for groomer in groomers:
        if not is_thin_listing(groomer):
            lastmod = groomer.get("last_modified", groomer.get("date_added", today)) or today
            entries.append((f"{config.SITE_URL}/groomer/{groomer['slug']}", "0.6", lastmod))

    for post in posts:
        if post.get("slug"):
            post_lastmod = post.get("modified_date", post.get("publish_date", today)) or today
            if post_lastmod > today:
                post_lastmod = today
            entries.append((f"{config.SITE_URL}/blog/{post['slug']}", "0.8", post_lastmod))

    data_by_url = {}
    for groomer in groomers:
        if not is_thin_listing(groomer):
            data_by_url[f"{config.SITE_URL}/groomer/{groomer['slug']}"] = groomer
    for post in posts:
        if post.get("slug"):
            data_by_url[f"{config.SITE_URL}/blog/{post['slug']}"] = post
    lastmods = _stable_lastmods(entries, data_by_url, today)

    sitemap = '<?xml version="1.0" encoding="UTF-8"?>\n'
    sitemap += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    for url, priority, lastmod in entries:
        sitemap += f"  <url><loc>{url}</loc><lastmod>{lastmods.get(url, lastmod)}</lastmod><priority>{priority}</priority></url>\n"
    sitemap += "</urlset>"

    output_path = config.OUTPUT_DIR / "sitemap.xml"
    output_path.write_text(sitemap)
    print("Built: sitemap.xml")


def build_robots():
    """Generate robots.txt."""
    robots = f"""User-agent: *
Allow: /

Sitemap: {config.SITE_URL}/sitemap.xml
"""
    output_path = config.OUTPUT_DIR / "robots.txt"
    output_path.write_text(robots)
    print("Built: robots.txt")


def build_netlify_redirects(groomers, posts, breeds=None, public_listing_pool=None):
    """Generate dist/_redirects with one literal 301 per .html page.

    Placeholder interpolation in Netlify redirects (both netlify.toml and
    _redirects file) was producing literal `:slug` strings instead of the
    captured value. Generating one explicit rule per page sidesteps the
    issue entirely. Netlify's documented limit is 100,000 rules per file;
    we're well under that.

    The trailing `!` is the force flag (redirect even if the file exists).
    """
    lines = [
        "# .html → extensionless 301 redirects (canonical URL cleanup)",
        "# Generated by build.py — do not edit by hand.",
        "",
        "# Static pages",
    ]

    static_pages = [
        "about", "contact", "privacy", "terms", "disclaimer",
        "editorial-standards", "submit", "blog",
    ]
    if breeds:
        static_pages.append("dog-grooming-guide")

    for name in static_pages:
        lines.append(f"/{name}.html  /{name}  301!")

    lines += ["", "# State pages"]
    for state in config.US_STATES:
        slug = state["slug"]
        lines.append(f"/state/{slug}.html  /state/{slug}  301!")

    lines += ["", "# City pages"]
    city_groups = group_by_city(public_listing_pool or groomers)
    for (state_slug, _city, city_slug), city_groomers in sorted(city_groups.items()):
        if len(city_groomers) >= 2:
            lines.append(
                f"/state/{state_slug}/{city_slug}.html  /state/{state_slug}/{city_slug}  301!"
            )

    lines += ["", "# Category pages"]
    for category in config.CATEGORIES:
        slug = category["slug"]
        lines.append(f"/category/{slug}.html  /category/{slug}  301!")

    lines += ["", "# Legacy category aliases"]
    lines.append("/category/mobile  /category/mobile-grooming  301!")

    lines += ["", "# Groomer pages"]
    gone = []
    for groomer in groomers:
        slug = groomer.get("slug")
        if not slug:
            continue
        if is_thin_listing(groomer):
            gone.append(slug)
            continue
        lines.append(f"/groomer/{slug}.html  /groomer/{slug}  301!")

    # Listings that fail the quality gate have no page (see build_groomer_pages).
    # 410 Gone is the correct signal: the content is permanently withdrawn, not moved.
    # A 301 to the state hub would be an "irrelevant or misleading" destination under
    # AdSense P2's navigation clause, since that hub does not contain the business.
    if gone:
        lines += ["", "# Withdrawn listings — 410 Gone (failed the content-quality gate)"]
        for slug in gone:
            lines.append(f"/groomer/{slug}  /410.html  410!")
            lines.append(f"/groomer/{slug}.html  /410.html  410!")

    lines += ["", "# Blog posts"]
    for post in posts:
        slug = post.get("slug")
        if slug:
            lines.append(f"/blog/{slug}.html  /blog/{slug}  301!")

    output_path = config.OUTPUT_DIR / "_redirects"
    output_path.write_text("\n".join(lines) + "\n")
    print(f"Built: _redirects ({len(lines)} lines)")


def build_gone_page(env):
    """Render /410.html — the body served for withdrawn listing URLs.

    Netlify serves this document with a 410 status via the rules in _redirects.
    It is noindexed: it is a status page, not content.
    """
    html = env.get_template("gone.html").render(
        page_title="Listing removed",
        meta_description="This dog groomer listing is no longer published.",
        request_path="/410",
        noindex=True,
        suppress_ads=True,
    )
    (config.OUTPUT_DIR / "410.html").write_text(html)
    print("Built: 410.html (withdrawn-listing page)")


def copy_ads_txt():
    """Copy ads.txt to output directory."""
    ads_txt_path = Path("ads.txt")
    if ads_txt_path.exists():
        shutil.copy(ads_txt_path, config.OUTPUT_DIR / "ads.txt")
        print("Built: ads.txt")


# Static pages
STATIC_PAGES = [
    {
        "template": "about.html",
        "output": "about.html",
        "title": "About Us",
        "description": "Learn about Dog Groomer Locator and our mission to help dog owners find trusted professional groomers.",
    },
    {
        "template": "privacy.html",
        "output": "privacy.html",
        "title": "Privacy Policy",
        "description": "Our privacy policy explains how we collect, use, and protect your information.",
    },
    {
        "template": "contact.html",
        "output": "contact.html",
        "title": "Contact Us",
        "description": "Get in touch with Dog Groomer Locator for questions, suggestions, or to submit a new listing.",
    },
    {
        "template": "terms.html",
        "output": "terms.html",
        "title": "Terms of Service",
        "description": "Terms and conditions for using Dog Groomer Locator.",
    },
    {
        "template": "success.html",
        "output": "success/index.html",
        "title": "Message Sent",
        "description": "Thank you for contacting us.",
    },
    {
        "template": "submit.html",
        "output": "submit.html",
        "title": "Submit a Dog Groomer",
        "description": "Submit a dog groomer or grooming salon to be added to our directory.",
    },
    {
        "template": "disclaimer.html",
        "output": "disclaimer.html",
        "title": "Disclaimer",
        "description": "Important disclaimers about the information provided on Dog Groomer Locator.",
    },
    {
        "template": "editorial-standards.html",
        "output": "editorial-standards.html",
        "title": "Editorial Standards",
        "description": "How we research, verify, and maintain the quality of information on Dog Groomer Locator.",
    },
]


def build_static_pages(env, groomers=None, all_groomers=None):
    """Build static informational pages.

    `groomers` is the published pool (drives "we list N groomers"). `all_groomers` is
    every record in the database, so the editorial-standards page can state honestly how
    many we hold vs. how many meet the publishing standard. Under AdSense P1 the
    methodology page is the site's primary evidence of curation, so these figures are
    computed from the build rather than hand-written, and cannot go stale.
    """
    total_count = len(groomers) if groomers else 0
    published_count = total_count
    held_count = len(all_groomers) if all_groomers else 0
    withheld_count = max(held_count - published_count, 0)
    for page in STATIC_PAGES:
        template = env.get_template(page["template"])
        # Canonical path is extensionless (e.g. /about instead of /about.html).
        # success/index.html → /success/ for the canonical.
        output = page["output"]
        if output.endswith("/index.html"):
            canonical_path = "/" + output[: -len("index.html")]
        elif output.endswith(".html"):
            canonical_path = "/" + output[: -len(".html")]
        else:
            canonical_path = "/" + output
        html = template.render(
            page_title=f"{page['title']} - {config.SITE_NAME}",
            meta_description=page["description"],
            request_path=canonical_path,
            total_count=total_count,
            published_count=published_count,
            held_count=held_count,
            withheld_count=withheld_count,
            data_checked_on=config.SITE_DATA_CHECKED_ON,
            noindex=page["output"].startswith("success/"),
            suppress_ads=True,
        )
        output_path = config.OUTPUT_DIR / page["output"]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(html)
        print(f"Built: {page['output']}")


def load_breed_guide():
    """Load breed grooming guide data from xlsx, falling back to JSON for CI/Netlify."""
    import json as _json
    xlsx_path = config.DATA_DIR / "dog_grooming_master.xlsx"
    json_path = config.DATA_DIR / "breed_grooming_guide.json"
    if not xlsx_path.exists():
        if json_path.exists():
            with open(json_path) as f:
                breeds = _json.load(f)
            breeds.sort(key=lambda b: b.get("name", ""))
            print(f"Loaded {len(breeds)} breeds from breed_grooming_guide.json (xlsx not found).")
            return breeds
        print("  Breed guide data not found, skipping.")
        return []

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    breeds = []
    for row in rows[1:]:
        breed = {
            "name": (row[0] or "").strip(),
            "coat_type": (row[1] or "").strip(),
            "shedding_level": (row[2] or "").strip(),
            "grooming_frequency": (row[3] or "").strip(),
            "best_grooming": (row[4] or "").strip(),
            "why": (row[5] or "").strip(),
            "special_care": (row[6] or "").strip(),
            "professional_grooming": (row[7] or "").strip(),
            "bathing_frequency": (row[8] or "").strip(),
            "hypoallergenic": (row[9] or "").strip(),
        }
        if breed["name"]:
            breeds.append(breed)

    breeds.sort(key=lambda b: b["name"])
    print(f"Loaded {len(breeds)} breeds from grooming guide.")
    return breeds


def link_breeds_to_posts(breeds, posts):
    """Point each breed card at its dedicated post, where one exists.

    Ten breeds (poodle, shih tzu, golden retriever, husky, yorkie, ...) are
    covered by BOTH /dog-grooming-guide and a dedicated blog post, so the two
    pages competed for the same query and Google often ranked neither well.
    Linking the guide's card to the post makes the relationship hub -> spoke
    instead of duplicate -> duplicate, and passes the guide's authority to the
    deeper page. Matching is on the breed name appearing in the post slug.
    """
    if not breeds or not posts:
        return
    for breed in breeds:
        name = str(breed.get("name", "")).lower()
        key = re.sub(r"[^a-z]+", "-", name).strip("-")
        if not key:
            continue
        breed["post_slug"] = ""
        breed["post_title"] = ""
        for post in posts:
            slug = str(post.get("slug", ""))
            if key and key in slug:
                breed["post_slug"] = slug
                breed["post_title"] = post.get("title", "")
                break


def build_breed_guide(env, breeds, posts=None):
    """Build the breed grooming guide page."""
    if not breeds:
        print("Skipped: breed guide (no data)")
        return

    link_breeds_to_posts(breeds, posts or [])
    template = env.get_template("breed-guide.html")

    # Group by coat type and shedding level for the template
    coat_types = sorted(set(b["coat_type"] for b in breeds))
    shedding_levels = sorted(set(b["shedding_level"] for b in breeds if b["shedding_level"]))

    html = template.render(
        breeds=breeds,
        coat_types=coat_types,
        shedding_levels=shedding_levels,
        page_title=f"Dog Breed Grooming Guide - {config.SITE_NAME}",
        meta_description=f"Complete grooming guide for {len(breeds)} dog breeds organized by coat type. Learn grooming routines, shedding levels, bathing frequency, and special care needs for every coat type.",
        request_path="/dog-grooming-guide",
    )

    output_path = config.OUTPUT_DIR / "dog-grooming-guide.html"
    output_path.write_text(html)
    print(f"Built: dog-grooming-guide.html ({len(breeds)} breeds)")


def download_groomer_images(groomers):
    """Download external photo URLs locally to avoid Google Places URL expiry."""
    images_dir = config.OUTPUT_DIR / "static" / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    source_images_dir = config.STATIC_DIR / "images"
    source_images_dir.mkdir(parents=True, exist_ok=True)

    downloaded = 0
    skipped = 0
    failed = 0

    headers = {
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://www.google.com/",
        "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
    }

    for groomer in groomers:
        url = groomer.get("photo_url", "").strip()
        slug = groomer.get("slug", "").strip()

        if not url or not slug:
            skipped += 1
            continue

        if url.startswith("/static/"):
            skipped += 1
            continue

        local_path = images_dir / f"{slug}.jpg"

        if local_path.exists():
            groomer["photo_url"] = f"/static/images/{slug}.jpg"
            skipped += 1
            continue

        try:
            response = requests.get(url, headers=headers, timeout=15, stream=True)
            if response.status_code == 200:
                with open(local_path, "wb") as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                source_path = source_images_dir / f"{slug}.jpg"
                if not source_path.exists():
                    shutil.copy2(local_path, source_path)
                groomer["photo_url"] = f"/static/images/{slug}.jpg"
                downloaded += 1
            else:
                print(f"  Image {response.status_code}: {groomer.get('name', slug)} — clearing photo_url in Airtable")
                groomer["photo_url"] = ""
                clear_airtable_photo_url(groomer.get("_airtable_id", ""))
                failed += 1
        except Exception as e:
            print(f"  Image error for {groomer.get('name', slug)}: {e} — clearing photo_url")
            groomer["photo_url"] = ""
            failed += 1

    print(f"Images: {downloaded} downloaded, {skipped} skipped/cached, {failed} failed")


def validate_build(groomers):
    """Build-time content validation — warn on data quality issues."""
    warnings = 0

    # Check for duplicate meta descriptions (using generated meta descriptions)
    descriptions = {}
    for g in groomers:
        meta_desc = generate_dynamic_meta_description("groomer", {
            "description": g.get("description", ""),
            "name": g.get("name", ""),
            "city": g.get("city", ""),
            "state": g.get("state", ""),
        })
        if meta_desc and meta_desc in descriptions:
            print(f"  WARNING: Duplicate meta description between '{g['name']}' and '{descriptions[meta_desc]}'")
            warnings += 1
        elif meta_desc:
            descriptions[meta_desc] = g["name"]

    # Check for thin descriptions
    thin_count = sum(1 for g in groomers if len(str(g.get("description", "")).strip()) < config.MIN_DESCRIPTION_LENGTH)
    if thin_count:
        print(f"  WARNING: {thin_count} groomers have descriptions under {config.MIN_DESCRIPTION_LENGTH} chars")

    # Check for duplicate slugs
    slugs = {}
    for g in groomers:
        s = g.get("slug", "")
        if s in slugs:
            print(f"  WARNING: Duplicate slug '{s}' between '{g['name']}' and '{slugs[s]}'")
            warnings += 1
        else:
            slugs[s] = g["name"]

    if warnings == 0:
        print("  Content validation passed — no issues found.")
    else:
        print(f"  Content validation completed with {warnings} warning(s).")


def main():
    """Main build process."""
    print(f"\n{'='*50}")
    print(f"Building {config.SITE_NAME}")
    print(f"{'='*50}\n")

    print("Setting up output directory...")
    setup_output_directory()

    print("\nFetching groomers...")
    groomers = get_groomers()
    annotate_listing_quality(groomers)
    public_listing_pool = public_groomers(groomers)
    print(f"Quality gate: {len(public_listing_pool)} public listings, {len(groomers) - len(public_listing_pool)} under review")

    print("\nDownloading groomer images...")
    download_groomer_images(groomers)

    print("\nValidating content...")
    validate_build(groomers)

    print("\nFetching blog posts...")
    posts = fetch_blog_posts()

    print("\nLoading breed guide...")
    breeds = load_breed_guide()

    env = create_jinja_env()

    print("\nBuilding pages...")
    build_homepage(env, public_listing_pool, posts)
    build_state_pages(env, public_listing_pool)
    build_city_pages(env, public_listing_pool)
    build_groomer_pages(env, groomers, public_listing_pool)
    build_category_pages(env, public_listing_pool)
    build_static_pages(env, public_listing_pool, all_groomers=groomers)
    build_blog_page(env, posts)
    build_post_pages(env, posts)
    build_breed_guide(env, breeds, posts)

    print("\nBuilding SEO files...")
    build_sitemap(public_listing_pool, posts, breeds)
    build_robots()
    build_netlify_redirects(groomers, posts, breeds, public_listing_pool)
    build_gone_page(env)
    copy_ads_txt()
    build_search_index(public_listing_pool)

    print(f"\n{'='*50}")
    print(f"Build complete! Output in: {config.OUTPUT_DIR}")
    print(f"{'='*50}")
    print(f"\nTo preview locally:")
    print(f"  cd {config.OUTPUT_DIR}")
    print(f"  python3 -m http.server 8000")
    print(f"  Open http://localhost:8000")


if __name__ == "__main__":
    main()
